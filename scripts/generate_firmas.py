#!/usr/bin/env python3
"""Genera 02-firmas/<slug>/firma.html desde scripts/datos.xlsx y scripts/firma.html (plantilla)."""

from __future__ import annotations

import html
import sys
from pathlib import Path
from urllib.parse import urlparse

try:
    import openpyxl
except ImportError:
    print("Instala openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPTS = Path(__file__).resolve().parent
ROOT = SCRIPTS.parent
FIRMAS_OUT = ROOT / "02-firmas"
DATOS_XLSX_DEFAULT = SCRIPTS / "datos.xlsx"
TEMPLATE_DEFAULT = SCRIPTS / "02-firmas" / "kenji-kawaida" / "firma.html"
# Fijamos la rama para evitar caché/resolve ambiguo (master vs main) en jsDelivr.
BASE = "https://cdn.jsdelivr.net/gh/kenjikv/stc-signature@main"

SLUG_BY_NAME: dict[str, str] = {
    "Rolando Kerlin Ruiz Justiniano": "rolando-ruiz",
    "Edgar Jaldin Torrico": "edgar-jaldin",
    "Ernesto Soto Roca": "ernesto-soto",
    "Edwin Calle Terrazas": "edwin-calle",
    "Kenji Kawaida Villegas": "kenji-kawaida",
    "Jorge Bergman Mostajo Pedraza": "jorge-mostajo",
    "Alcides Yohacin Leaños Rodríguez": "alcides-leanos",
    "Paul Fernando Grimaldo Bravo": "paul-grimaldo",
    "Juan Carlos Peinado Pereira": "juan-carlos-peinado",
}


def is_url(v: object) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    if not s or s == "-":
        return False
    if s.upper() == "X":
        return False
    return s.startswith("http://") or s.startswith("https://")


def norm_email(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s == "-":
        return None
    return s


def fmt_phone(v: object) -> str | None:
    if v is None or v == "-":
        return None
    if isinstance(v, float):
        s = str(int(v))
    else:
        s = str(v).strip()
    if not s or s == "-":
        return None
    digits = "".join(c for c in s if c.isdigit())
    if not digits:
        return None
    if len(digits) == 8:
        return f"+591 {digits[:3]} {digits[3:]}"
    return f"+{digits}"


def display_host(url: str) -> str:
    try:
        p = urlparse(url)
        host = (p.netloc or "").lower()
        if host.startswith("www."):
            host = host[4:]
        return host or url
    except Exception:
        return url


def social_btn(href: str, title: str, icon: str, alt: str) -> str:
    href_e = html.escape(href, quote=True)
    src = f"{BASE}/01-iconos/{icon}"
    return f"""          <a href="{href_e}" class="sbtn" title="{html.escape(title)}" style="display:inline-block;vertical-align:middle;text-decoration:none;margin-right:4px;border:0;">
            <table role="presentation" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;mso-table-lspace:0pt;mso-table-rspace:0pt;">
              <tr>
                <td align="center" valign="middle" width="26" height="26" style="width:26px;height:26px;padding:0;line-height:0;font-size:0;border:1px solid #D7DCE8;border-radius:5px;">
                  <img src="{html.escape(src, quote=True)}" width="12" height="12" alt="{html.escape(alt)}" style="display:block;width:12px;height:12px;margin:0;border:0;outline:none;" />
                </td>
              </tr>
            </table>
          </a>"""


def build_kv_rows(
    web_propia: str | None,
    telefono: str | None,
    web_comunidad: str | None,
    correo: str | None,
) -> str:
    kv_parts: list[str] = []
    if is_url(web_propia):
        u = str(web_propia).strip()
        label = html.escape(display_host(u))
        kv_parts.append(
            f'          <div><span class="k" style="color:#9BA5BD;display:inline-block;width:56px;">web</span><a href="{html.escape(u, quote=True)}" style="color:#0B2A5B;text-decoration:none;">{label}</a></div>'
        )
    if telefono:
        kv_parts.append(
            f'          <div><span class="k" style="color:#9BA5BD;display:inline-block;width:56px;">tel</span>{html.escape(telefono)}</div>'
        )
    if is_url(web_comunidad):
        u = str(web_comunidad).strip()
        label = html.escape(display_host(u))
        kv_parts.append(
            f'          <div><span class="k" style="color:#9BA5BD;display:inline-block;width:56px;">stc</span><a href="{html.escape(u, quote=True)}" style="color:#32405A;text-decoration:none;">{label}</a></div>'
        )
    if correo:
        mailto = f"mailto:{correo}"
        kv_parts.append(
            f'          <div><span class="k" style="color:#9BA5BD;display:inline-block;width:56px;">email</span><a href="{html.escape(mailto, quote=True)}" style="color:#32405A;text-decoration:none;">{html.escape(correo)}</a></div>'
        )
    if kv_parts:
        return "\n".join(kv_parts)
    return '          <div><span class="k" style="color:#9BA5BD;display:inline-block;width:56px;">stc</span><a href="https://stc.soeuagrm.edu.bo/" style="color:#32405A;text-decoration:none;">stc.soeuagrm.edu.bo</a></div>'


def build_social_buttons(
    linkedin: str | None,
    x_url: str | None,
    facebook: str | None,
    instagram: str | None,
    medium: str | None,
) -> str:
    socials: list[str] = []
    if is_url(linkedin):
        socials.append(social_btn(str(linkedin).strip(), "LinkedIn", "ic_linkedin.png", "LinkedIn"))
    if is_url(x_url):
        socials.append(social_btn(str(x_url).strip(), "X", "ic_x.png", "X"))
    if is_url(facebook):
        socials.append(social_btn(str(facebook).strip(), "Facebook", "ic_facebook.png", "Facebook"))
    if is_url(instagram):
        socials.append(social_btn(str(instagram).strip(), "Instagram", "ic_instagram.png", "Instagram"))
    if is_url(medium):
        socials.append(social_btn(str(medium).strip(), "Medium", "ic_medium.png", "Medium"))
    if not socials:
        return ""
    return "\n" + "\n".join(socials) + "\n        "


def render_firma(
    template: str,
    *,
    full_name: str,
    cargo: str,
    slug: str,
    linkedin: str | None,
    facebook: str | None,
    x_url: str | None,
    instagram: str | None,
    medium: str | None,
    telefono: str | None,
    correo: str | None,
    web_propia: str | None,
    web_comunidad: str | None,
) -> str:
    cargo = (cargo or "").strip()
    photo_src = f"{BASE}/02-firmas/{slug}/foto.jpg"
    logo_src = f"{BASE}/01-iconos/logo.png"
    kv_rows = build_kv_rows(web_propia, telefono, web_comunidad, correo)
    social_buttons = build_social_buttons(linkedin, x_url, facebook, instagram, medium)

    subs = {
        "{{PHOTO_SRC}}": html.escape(photo_src, quote=True),
        "{{LOGO_SRC}}": html.escape(logo_src, quote=True),
        "{{FULL_NAME}}": html.escape(full_name),
        "{{CARGO}}": html.escape(cargo),
        "{{KV_ROWS}}": kv_rows,
        "{{SOCIAL_BUTTONS}}": social_buttons,
    }
    out = template
    for key, val in subs.items():
        if key not in out:
            print(f"Aviso: la plantilla no contiene el marcador {key}", file=sys.stderr)
        out = out.replace(key, val)
    out = out.replace(
        "<!-- STC Signature · plantilla · no editar salidas en 02-firmas a mano: usar generate_firmas.py -->",
        "<!-- STC Signature · generada con generate_firmas.py · no editar a mano -->",
        1,
    )
    return out


def load_template(path: Path) -> str:
    if not path.is_file():
        print(f"No existe la plantilla: {path}", file=sys.stderr)
        sys.exit(1)
    t = path.read_text(encoding="utf-8")
    required = ("{{PHOTO_SRC}}", "{{LOGO_SRC}}", "{{FULL_NAME}}", "{{CARGO}}", "{{KV_ROWS}}", "{{SOCIAL_BUTTONS}}")
    for m in required:
        if m not in t:
            print(f"Error: la plantilla debe incluir el marcador {m}", file=sys.stderr)
            sys.exit(1)
    return t


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DATOS_XLSX_DEFAULT
    tpl_path = Path(sys.argv[2]) if len(sys.argv) > 2 else TEMPLATE_DEFAULT

    if not xlsx.is_file():
        print(f"No existe el Excel: {xlsx}", file=sys.stderr)
        sys.exit(1)

    template = load_template(tpl_path)
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.strip() if isinstance(c, str) else c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    def col(row: tuple, name: str) -> object:
        i = idx.get(name)
        if i is None:
            return None
        return row[i] if i < len(row) else None

    written = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        slug = SLUG_BY_NAME.get(name)
        if not slug:
            print(f"Aviso: sin carpeta/slug para «{name}», se omite.", file=sys.stderr)
            continue

        cargo = col(row, "Cargo")
        linkedin = col(row, "Linkedin")
        facebook = col(row, "Facebook")
        x_url = col(row, "X")
        instagram = col(row, "Instagram")
        medium = col(row, "Medium/Otra red social")
        tel_raw = col(row, "Telefono")
        correo = norm_email(col(row, "Correo electronico"))
        web_p = col(row, "Página web propia")
        web_c = col(row, "Página web de la comunidad")

        out = FIRMAS_OUT / slug / "firma.html"
        out.parent.mkdir(parents=True, exist_ok=True)
        body = render_firma(
            template,
            full_name=name,
            cargo=str(cargo or "").strip(),
            slug=slug,
            linkedin=str(linkedin).strip() if linkedin else None,
            facebook=str(facebook).strip() if facebook else None,
            x_url=str(x_url).strip() if x_url else None,
            instagram=str(instagram).strip() if instagram else None,
            medium=str(medium).strip() if medium else None,
            telefono=fmt_phone(tel_raw),
            correo=correo,
            web_propia=str(web_p).strip() if web_p else None,
            web_comunidad=str(web_c).strip() if web_c else None,
        )
        out.write_text(body, encoding="utf-8")
        print(out)
        written += 1

    wb.close()
    print(f"Listo: {written} firmas (plantilla: {tpl_path}).", file=sys.stderr)


if __name__ == "__main__":
    main()
